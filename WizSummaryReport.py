import os
import pandas as pd
from openpyxl import Workbook
from utilities.logger_master import logger, log_function_entry_exit


@log_function_entry_exit(logger)
class WizSummaryReport:
    def __init__(self, root_folder, output_file_path):
        logger.info("Initializing WizSummaryReport")
        self.root_folder = root_folder if isinstance(root_folder, list) else [root_folder]
        self.output_file_path = output_file_path
        self.metrics = {}
        self.row_val_map = {
            'Total assets': 0,
            'Unique Vulnerable assets older than 45 days': 0,
            'Unique Exploitable assets older than 45 days': 0,
            'Unique CISA KEV assets older than 45 days': 0,
        }
        # Update below mapping
        # sheet name : [list of metrics]
        self.metric_grouping = {
            'HIS': ['DataLake', 'OtisOne', 'NE CRM'],
            'Stand': ['ABC', 'XFZ', '1G3']
        }

    
    def extract_metrics(self):
        logger.info(f'Starting to extract metrics from folder(s): {self.root_folder}')
        for folder in self.root_folder:
            for subdir, dirs, files in os.walk(folder):
                for file in files:
                    if file.endswith(".xlsx") and '_count' in file:
                        category = file.split('_count')[0]
                        file_path = os.path.join(subdir, file)
                        logger.info(f'Category: {category}, Processing file: {file_path}')
                        try:
                            self._process_excel_file(file_path, category)
                        except Exception as e:
                            logger.error(f"Error processing file {file_path}: {e}")

    def _process_excel_file(self, file_path, category):
        logger.info(f'Extracting data from file: {file_path}')
        excel_file = pd.ExcelFile(file_path)
        sheet_name = ''
        if 'count' in excel_file.sheet_names:
            sheet_name = 'count'
        elif 'counts' in excel_file.sheet_names:
            sheet_name = 'counts'

        if sheet_name != '':
            logger.info(f"'{sheet_name}' sheet found in {file_path}")
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            overall_category = False
            for i, row in df.iterrows():
                col_a_summary = row['Summary Description']
                col_b_value = row['Summary Values']
                if sheet_name == 'count':
                    if col_a_summary in self.row_val_map and pd.notna(col_b_value):
                        logger.info(f"Updating {col_a_summary} for category {category} with value {col_b_value}")
                        self.metrics[category][col_a_summary] = int(float(col_b_value))

                elif sheet_name == 'counts':
                    if str(col_a_summary) == 'Category' and str(col_b_value) == 'Overall':
                        overall_category = True
                        logger.info(f"Found category: {overall_category}")
                        self.metrics[category] = {key: 0 for key in self.row_val_map}
                    elif overall_category and col_a_summary in self.row_val_map and pd.notna(col_b_value):
                        logger.info(f"Updating {col_a_summary} for category {category} with value {col_b_value}")
                        self.metrics[category][col_a_summary] = int(float(col_b_value))
        else:
            logger.warning(f"'count' sheet not found in {file_path}")

    def create_summary_workbook(self):
        logger.info('Creating summary workbook')
        os.makedirs(os.path.dirname(self.output_file_path), exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, metric_list in self.metric_grouping.items():
            logger.info(f'Creating sheet for: {sheet_name}')
            sheet = wb.create_sheet(title=sheet_name)
            self._populate_sheet(sheet, metric_list)

        if len(self.metrics) > 0:
            logger.info('Creating sheet for remaining metrics (other)')
            sheet = wb.create_sheet(title='other')
            self._populate_sheet(sheet, list(self.metrics.keys()))

        wb.save(self.output_file_path)
        logger.info(f'Summary workbook saved to: {self.output_file_path}')

    def _populate_sheet(self, sheet, metric_list):
        logger.info(f"Populating sheet {sheet.title}")
        self.sheet_summary = {key: 0 for key in self.row_val_map}

        for metric in metric_list:
            if metric in self.metrics:
                logger.info(f"Adding data for metric: {metric}")
                sheet.append([metric, ''])
                for key, val in self.metrics[metric].items():
                    sheet.append([key, val])
                    self.sheet_summary[key] += val

                sheet.append(['', ''])
                sheet.append(['', ''])

                del self.metrics[metric]

        sheet.append(['Overall Metrics', ''])
        for key, val in self.sheet_summary.items():
            sheet.append([key, val])

        sheet.append(['', ''])
        sheet.append(['', ''])

        total_assets = self.sheet_summary.get('Total assets', 0)
        for key, val in self.sheet_summary.items():
            if key != 'Total assets':
                percentage = round(val / total_assets, 3) if total_assets > 0 else 'NA'
                logger.info(f"Calculating percentage for {key}: {percentage}")
                sheet.append([f'{key} percentage', percentage])

    def run(self):
        logger.info("Running WizSummaryReport - START")
        self.extract_metrics()
        self.create_summary_workbook()
        logger.info("Running WizSummaryReport - END")


# update folder path to excel files
input_folder_path = 'input'
output_file_path = "output/metrics_summary.xlsx"
wiz_summary = WizSummaryReport(input_folder_path, output_file_path)
wiz_summary.run()

